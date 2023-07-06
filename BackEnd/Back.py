from Clases import Excel
from openpyxl import load_workbook
from Utils import HelperGui


class BackClass:

    def __init__(self):
        self.excel = Excel.ExcelClass()

    def reinicio(self):
        self.excel.getLibro().close()
        self.excel.reinicio()

    def maximoFilas(self):
        return self.excel.maximoFilas()

    def maximoColumnas(self):
        return self.excel.maximoColumnas()

    # Funcion que recibe la ruta y el nombre de la hoja de un excel y lo abre
    def abrirExcel(self, path, nombrehoja=None):
        # Abrimos el excel
        libro = load_workbook(filename=path, read_only=False)
        # Abrimos la hoja, si no tiene nombre la abrimos bajo el nombre default
        if nombrehoja is None or nombrehoja == "":
            nombrehoja = "Hoja1"

        hoja = libro[nombrehoja]
        self.excel = Excel.ExcelClass(libro, hoja, path)

    """
            FUNCIONES PARA EL INDICE
    """
    # Funcion que Crea el un indice de las filas verticales segun los datos que existen en el excel
    def crearIndice(self):
        tamFilas = 5
        if self.excel.maximoFilas() < 5:
            tamFilas = self.excel.maximoFilas()

        tamColums = self.excel.maximoColumnas()
        for columna in range(tamColums):
                tam = 0
                tipo = "string"
                if tamFilas == 1:
                    tipo = "Indeterminado"
                tamP = 0
                tamM = 0
                tamG = 0
                nones = 0
                for linea in range(tamFilas):
                    celda = self.excel.hoja.cell(row=linea + 2, column=columna + 1)
                    if celda.data_type == "f":
                        tipo = "Funcion"
                        tam = 1/5
                        break
                    elif celda.value == "true" or celda.value == "false":
                        tipo = "Booleano"
                        tam = 1 / 7
                        break
                    elif isinstance(celda.value, int):
                        tipo = "Integer"
                        tam = 1 / 5
                        break
                    else:
                        if celda.value is not None and celda.value != "":
                            if len(celda.value) > 15 and len(celda.value) <= 30:
                                tamM = tamM + 1
                            elif len(celda.value) > 30:
                                tamG = tamG + 1
                            else:
                                tamP = tamP + 1
                        else:
                            nones = nones + 1
                if tamP > tamM and tamP > tamG and tipo == "string":
                    tipo = "short string"
                    tam = 1 / 5
                elif tamM >= tamP and tamM > tamG and tipo == "string":
                    tipo = "string"
                    tam = 1 / 3
                elif tamG >= tamP and tamG >= tamP and tipo == "string":
                    tipo = "long string"
                    tam = 1 / 3
                color = self.obtenerColor(columna + 1)
                nombre = self.obtenerNombre(columna + 1)
                if nones != tamFilas:
                    self.excel.setIntoIndice(columna + 1, tipo, color, nombre, tam)

    def getIndice(self):
        return self.excel.getIndice()

    """
            FUNCIONES PARA OBTENER VALORES
    """
    #Funcion que recupera el valor el cual pondremos como nombre
    def obtenerNombre(self, columna):
        celda = self.excel.hoja.cell(row=1, column=columna)
        return celda.value

    # Funcion que recupera el valor del color transformandolo a hexadecimal
    def obtenerColor(self, columna):
        celda = self.excel.hoja.cell(row=1, column=columna)
        color = ""
        if celda.fill.start_color.index is not None:
            color = celda.fill.start_color.index
        else:
            color = '#FFFFFFFF'
        if color == '00000000' or color == 0:
            color = '#FFFFFFFF'
        return HelperGui.cambioHexARGB(color)


    """
            FUNCIONES DE LAS LINEAS
    """
    # Funcion que recupera una columna a partir de un dato
    def obtenerLinea(self, pos, dato):
        # Buscamos la posicion en la que esta el dato
        num = self.excel.buscarDato(pos, dato)
        if num != -1:  # si existe devolvemos los datos
            return num, self.excel.buscarColumna(num)
        return 1, None

    def anadirFila(self, linea):
        self.excel.insertarColumna(linea)

    def borrarFila(self, filas):

        posicion = self.excel.buscarDato(filas[0][0], filas[0][1])
        if posicion != -1:
            # buscamos que todos los valores esten
            for fila in filas:
                if not self.excel.getValorIndice(fila[0]).esFuncion() and not self.excel.getValorIndice(
                        fila[0]).esBooleano():
                    posicion2 = self.excel.buscarDato(fila[0], fila[1])
                    if posicion2 != posicion:
                        return -1

            self.excel.eliminarColumna(posicion)
            self.reorganizarFunciones(posicion)
        return 1

    def siguienteLinea(self, posicion):
        if posicion < self.excel.maximoFilas():
            return self.excel.buscarColumna(posicion + 1)

    def anteriorLinea(self, posicion):
        if posicion >= 1:
            return self.excel.buscarColumna(posicion - 1)

    def modificarFila(self, posicion, filas):
        if posicion <= self.excel.maximoFilas():
            self.excel.modificarColumna(filas, posicion)

    def ordenarFilas(self, numfila, orden):
        filas = self.maximoFilas()
        if filas > 2:
            # Obtener los datos de la hoja, excluyendo la primera fila
            datos = list(self.excel.getHoja().iter_rows(min_row=2, max_row=filas, values_only=True))
            # Ordenar los datos basándose en la primera columna
            if self.excel.getValorIndice(numfila + 1).esInteger():
                datos_ordenados = sorted(datos, key=lambda x: int(x[numfila]), reverse=orden)
            else:
                datos_ordenados = sorted(datos, key=lambda x: str(x[numfila]), reverse=orden)

            self.excel.insertarOrdenados(datos_ordenados)

            self.excel.eliminarAntiguos(filas)

    def reorganizarFunciones(self, posicion):
        maxcol = self.maximoColumnas()
        for posx in range(self.maximoColumnas()):
            maxfil = self.maximoFilas()
            if self.excel.getValorIndice(posx + 1).esFuncion():
                for posy in range(posicion, self.maximoFilas() + 1):
                    # obtenemos la formula de un valor por encima
                    dato = self.excel.cambiarFormula(self.excel.getHoja().cell(row=posy, column=posx + 1).value, -1)
                    # la introducimos en la nueva columna
                    self.excel.getHoja().cell(row=posy, column=posx + 1).value = dato
                self.excel.getLibro().save(self.excel.ruta)
